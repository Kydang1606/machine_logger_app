import streamlit as st
import pandas as pd
import plotly.express as px
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook

# ================================================
# 🛠️ CẤU HÌNH BAN ĐẦU
# ================================================
st.set_page_config(page_title="Machine Log App", layout="centered")
DATA_PATH = "data/Logs.xlsx"
SHEET_NAME = "Logs"

# ================================================
# 🧠 HÀM TẠO FILE EXCEL VÀ SHEET "Logs" NẾU CHƯA TỒN TẠI
# ================================================
def ensure_log_file():
    if not os.path.exists("data"):
        os.makedirs("data")

    if not os.path.exists(DATA_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append([
            "Date", "Start", "End", "Total (min)", "Total (hr)",
            "Machine", "Project", "Material", "Employee", "Description"
        ])
        wb.save(DATA_PATH)
    else:
        wb = load_workbook(DATA_PATH)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append([
                "Date", "Start", "End", "Total (min)", "Total (hr)",
                "Machine", "Project", "Material", "Employee", "Description"
            ])
            wb.save(DATA_PATH)

# ================================================
# ✅ GHI DỮ LIỆU VÀO FILE EXCEL
# ================================================
def log_to_excel(row_dict):
    df_new = pd.DataFrame([row_dict])

    with pd.ExcelWriter(DATA_PATH, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        reader = pd.read_excel(DATA_PATH, sheet_name=SHEET_NAME)
        df_all = pd.concat([reader, df_new], ignore_index=True)
        df_all.to_excel(writer, sheet_name=SHEET_NAME, index=False)

# ================================================
# 🚀 ỨNG DỤNG CHÍNH
# ================================================
def main():
    st.title("📝 Machine Work Logger")
    ensure_log_file()

    with st.form("log_form", clear_on_submit=True):
        col1, col2, col3 = st.columns(3)
        with col1:
            date = st.date_input("📅 Date", value=datetime.today())
            start_time = st.time_input("🕑 Start Time", value=datetime.strptime("08:00", "%H:%M").time())
        with col2:
            end_time = st.time_input("🕕 End Time", value=datetime.strptime("12:00", "%H:%M").time())
            machine = st.text_input("🏭 Machine", placeholder="e.g., CNC01")
        with col3:
            project = st.text_input("📁 Project Code", placeholder="e.g., D001")
            material = st.text_input("🧱 Material Type", placeholder="e.g., CFRP, GRP,...")

        employee = st.text_input("👤 Employee")
        description = st.text_area("📋 Work Description", height=100)

        submitted = st.form_submit_button("✅ Submit Entry")
        if submitted:
            try:
                start_dt = datetime.combine(date, start_time)
                end_dt = datetime.combine(date, end_time)
                if end_dt < start_dt:
                    end_dt += timedelta(days=1)

                total_minutes = int((end_dt - start_dt).total_seconds() // 60)
                total_hours = round(total_minutes / 60, 2)

                row = {
                    "Date": date,
                    "Start": start_time.strftime("%H:%M"),
                    "End": end_time.strftime("%H:%M"),
                    "Total (min)": total_minutes,
                    "Total (hr)": total_hours,
                    "Machine": machine,
                    "Project": project,
                    "Material": material,
                    "Employee": employee,
                    "Description": description
                }

                log_to_excel(row)
                st.success("✅ Entry logged successfully!")
                st.experimental_rerun()

            except Exception as e:
                st.error(f"❌ Error: {e}")

    st.markdown("---")

    # ========================================
    # 📋 HIỂN THỊ LOGS VÀ BỘ LỌC
    # ========================================
    st.header("📊 Logs Viewer")

    try:
        df_logs = pd.read_excel(DATA_PATH, sheet_name=SHEET_NAME)
        if df_logs.empty:
            st.info("Chưa có dữ liệu nào.")
            return

        filter_col = st.selectbox("🔍 Filter by", ["Project", "Machine"])
        options = df_logs[filter_col].dropna().unique()
        selected = st.multiselect(f"Select {filter_col}(s)", options, default=options)

        filtered_df = df_logs[df_logs[filter_col].isin(selected)]
        st.dataframe(filtered_df, use_container_width=True)

        # ========================================
        # 📈 BIỂU ĐỒ TỔNG GIỜ
        # ========================================
        st.subheader("📊 Total Hours Chart")
        chart_data = (
            filtered_df.groupby(filter_col)["Total (hr)"]
            .sum()
            .reset_index()
            .sort_values("Total (hr)", ascending=False)
        )

        fig = px.bar(
            chart_data,
            x=filter_col,
            y="Total (hr)",
            title=f"Total Hours by {filter_col}",
            text_auto=".2s",
            color=filter_col,
            height=400
        )
        st.plotly_chart(fig, use_container_width=True)

    except Exception as e:
        st.error(f"⚠️ Error reading logs: {e}")

# ================================================
# 🔁 CHẠY APP
# ================================================
if __name__ == "__main__":
    main()
